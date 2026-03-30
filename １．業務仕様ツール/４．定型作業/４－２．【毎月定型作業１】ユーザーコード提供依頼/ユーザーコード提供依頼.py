import os
import openpyxl
import sqlite3
import pandas as pd
import time
import pyautogui
# 実行ファイルのパスを取得
import os
import sys

# カレントパス取得、共通モジュールのパスを絶対パスで固定
#current_dir = os.path.dirname(os.path.abspath(__file__)) 
#common_folder_path = r'C:\ローカルツール\python\共通'

# 実行ファイルがあるディレクトリの絶対パスを取得→相対パスで共通フォルダを設定
# '../..' で 2 つ上のディレクトリへ移動
current_dir = os.path.dirname(os.path.abspath(__file__)) 
common_folder_path = os.path.join(current_dir, '..', '..', '..', '共通')

# sys.pathに共通フォルダを追加し、共通モジュールをインポートする
sys.path.append(common_folder_path)
from common_func import *
from typing import List

# ① c:\tmpディレクトリ配下のファイルをリネーム
def rename_files(directory):
    for filename in os.listdir(directory):
        if "ユーザーコード提供依頼_" in filename and filename.endswith(".xlsx"):
            old_path = os.path.join(directory, filename)
            new_filename = filename.replace(".xlsx", "_ユーザーコード追記.xlsx")
            new_path = os.path.join(directory, new_filename)
            os.rename(old_path, new_path)
            print(f"Renamed: {filename} -> {new_filename}")
            return new_path  # リネームしたファイルのパスを返す

# ② リネーム後のファイルを開く
def open_excel_file(file_path):
    return openpyxl.load_workbook(file_path)

# ③ SQLファイルを作成する
def create_sql_file(excel_file,current_dir):
    sheet = excel_file.active  # 1つめのシートを取得
#    sql_file_path = current_dir+'/sql/ユーザーコード提供依頼.sql'
    sql_file_path = current_dir+'/ユーザーコード提供依頼.sql'
    os.makedirs(os.path.dirname(sql_file_path), exist_ok=True)

    with open(sql_file_path, 'w', encoding='utf-8') as f:
        f.write("select distinct inquiry_num, USER_CD from tjfax151_customer_request a where ( inquiry_num in (\n")
        
        # A列の2行目から最後まで読み取り
        row = 2
        while sheet[f"A{row}"].value:

            inquiry_num = sheet[f"A{row}"].value

            if str(row).endswith('00'):
                f.write(f"'{inquiry_num}') or inquiry_num in ('',\n")
            else:
                f.write(f"'{inquiry_num}',\n")

            row += 1
        
        f.write("''))\nand CR_CTGR not in ('16')\norder by inquiry_num\n")
    
    print(f"SQLファイルが作成されました: {sql_file_path}")
    return sql_file_path

# ④ SQLファイルを実行する
def execute_sql_file():
    # SQLiteデータベースを仮定して実行
    conn = sqlite3.connect('database.db')  # 実際のDB接続に置き換える
    cursor = conn.cursor()

    with open('./sql/ユーザーコード提供依頼.sql', 'r', encoding='utf-8') as f:
        sql_query = f.read()

    cursor.execute(sql_query)
    conn.commit()
    print("SQLクエリが実行されました")

    # 結果をDataFrameに読み込む
    result = pd.read_sql_query(sql_query, conn)
    conn.close()
    
    return result

# ⑤ 実行結果をExcelに出力
def write_results_to_excel(original_file, result_df):
    # 元のファイルを再度開く
    file_path = original_file
    wb = openpyxl.load_workbook(file_path)

    # 新しいシートを追加
    sheet = wb.create_sheet('Sheet2')
    
    # DataFrameの内容をシートに書き込む
    for row in dataframe_to_rows(result_df, index=False, header=True):
        sheet.append(row)

    # 保存
    wb.save(file_path)
    print(f"結果をExcelに出力しました: {file_path}")


# ⑥ 「Sheet2」をアクティブにして開く
def open_excel_with_sheet2(file_path):

    wb = openpyxl.load_workbook(file_path)
    wb.active = wb['Sheet2']  # Sheet2をアクティブにする

    os.startfile(file_path)  # Excelを開く
 
    # アクティブシートを取得
    sheet = wb.active
    
    # B列（2列目）のセルを走査
    for row in sheet.iter_rows(min_col=2, max_col=2, min_row=1, max_row=sheet.max_row):
        for cell in row:
            # セルの値が「None」の場合、それを空文字に変換
            if cell.value == "None":
                cell.value = ""

    print("Excelファイルが開かれました")
    wb.save(file_path)




def rename_and_add_sheet(tmp_output_file_name, renamed_file):
    # tmp_output_file_name のエクセルファイルを開く
    tmp_wb = openpyxl.load_workbook(tmp_output_file_name)
    
    # tmp_output_file_name にある1つ目のシートを取得
    tmp_sheet = tmp_wb.active
    
    # シートを "Sheet2" にリネーム
    tmp_sheet.title = "Sheet2"
    
    # renamed_file のエクセルファイルを開く
    output_wb = openpyxl.load_workbook(renamed_file)
    
    # 新しいシートを作成し、tmp_sheetの内容をコピー
    new_sheet = output_wb.create_sheet("Sheet2")
    
    # tmp_sheetの内容を新しいシートにコピー
    for row in tmp_sheet.iter_rows():
        for cell in row:
            new_sheet[cell.coordinate].value = cell.value

    # output_file_name を保存
    output_wb.save(renamed_file)
    print(f"シート 'Sheet2' を {renamed_file} に追加しました。")

    wb = openpyxl.load_workbook(renamed_file)
    wb.active = wb["Sheet1"]
    wb.active = wb["Sheet2"]





# 実行の流れ
def main():
    # ① リネーム
    directory = r"c:\tmp"
    renamed_file = rename_files(directory)
    print(current_dir)

    # ② Excelファイルを開く
    excel_file = open_excel_file(renamed_file)
    
    # ③ SQLファイルを作成
    sql_file_path=create_sql_file(excel_file,current_dir)
    
    # ④ SQLファイルを実行
    # 入力引数のリスト
    input_content: List[str] = ["", "", "", "", "", "", "", "", "", ""]
    output_file_name=common_output_excel("ffm",sql_file_path,input_content)
    tmp_output_file_name=directory+'/'+common_get_filename_from_path(output_file_name)

    #print("output_file_name:"+output_file_name)
    #print ("tmp_output_file_name:"+tmp_output_file_name)

    common_move_file(output_file_name,tmp_output_file_name)

    #print ("tmp_output_file_name:"+tmp_output_file_name)
    #print("renamed_file:"+renamed_file)

    rename_and_add_sheet(tmp_output_file_name, renamed_file)

    # ⑥ 「Sheet2」をアクティブにして開く
    open_excel_with_sheet2(renamed_file)
    # ページダウンキーを押す 
    time.sleep(2) 
    pyautogui.hotkey('ctrl', 'pageup')
    pyautogui.hotkey('ctrl', 'pagedown')

    # 回答コメントを開く
    text_file_path = r"C:\Users\g10012013\OneDrive - Ricoh\ツール\１．コマンド実行\３．データ取得\DataStudio「問合せ番号」に対する「リコー専用項目」データ抽出の件\回答.txt"
    os.system(f"notepad {text_file_path}")
    

    

    

if __name__ == "__main__":
    main()
