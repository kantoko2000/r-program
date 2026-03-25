from openpyxl import load_workbook
# 共通関数インポート
import os
import sys

# カレントパス取得、共通モジュールのパスを絶対パスで固定
current_dir = os.path.dirname(os.path.abspath(__file__)) 
common_folder_path = r'C:\ローカルツール\python\共通'

# sys.pathに共通フォルダを追加し、共通モジュールをインポートする
sys.path.append(common_folder_path)
from common_func import *
import win32com.client as win32
import xlwings as xw

# 書面審査報告書格納ディレクトリ
syomen_report_dir = r"\\z2stzsmbxx0199.file.core.windows.net\z2sh16000zsmb199\share\FFM_書面審査\作業履歴\★定型作業\報告書送付\\"

# セルの表示形式を変換する関数
def process_excel_columns(file_path):

    # Excel アプリケーションを起動
    excel = win32.Dispatch('Excel.Application')
    # Excelウィンドウを非表示にする（Trueにすると表示される）
    excel.Visible = False

    # Excel ファイルを開く
    workbook = excel.Workbooks.Open(file_path)

    # 1つ目のシートを選択
    sheet = workbook.Sheets(1)

    # 「AS列」、「AV列」、「AW列」から「EL列」まで順に処理
    #columns_to_process = ['AS', 'AV', 'AW', 'AX', 'AY', 'AZ', 'BA', 'BB', 'BC', 'BD', 'BE', 'BF', 'BG', 'BH', 'BI', 'BJ', 'BK', 'BL', 'BM', 'BN', 'BO', 'BP', 'BQ', 'BR', 'BS', 'BT', 'BU', 'BV', 'BW', 'BX', 'BY', 'BZ', 'CA', 'CB', 'CC', 'CD', 'CE', 'CF', 'CG', 'CH', 'CI', 'CJ', 'CK', 'CL', 'CM', 'CN', 'CO', 'CP', 'CQ', 'CR', 'CS', 'CT', 'CU', 'CV', 'CW', 'CX', 'CY', 'CZ', 'DA', 'DB', 'DC', 'DD', 'DE', 'DF', 'DG', 'DH', 'DI', 'DJ', 'DK', 'DL', 'DM', 'DN', 'DO', 'DP', 'DQ', 'DR', 'DS', 'DT', 'DU', 'DV', 'DW', 'DX', 'DY', 'DZ', 'EA', 'EB', 'EC', 'ED', 'EE', 'EF', 'EG', 'EH', 'EI', 'EJ', 'EK', 'EL']
    # 右記の列は変換すると表示に支障があるので、変換しない「BD、BZ、DI、EF、日時が壊れる・・・DP列、CQ列、BW列」
    columns_to_process = ['AS', 'AV', 'AW', 'AX', 'AY', 'AZ', 'BA', 'BB', 'BC', 'BE', 'BF', 'BG', 'BH', 'BI', 'BJ', 'BK', 'BL', 'BM', 'BN', 'BO', 'BP', 'BQ', 'BR', 'BS', 'BT', 'BU', 'BV', 'BX', 'BY',  'CA', 'CB', 'CC', 'CD', 'CE', 'CF', 'CG', 'CH', 'CI', 'CJ', 'CK', 'CL', 'CM', 'CN', 'CO', 'CP', 'CR', 'CS', 'CT', 'CU', 'CV', 'CW', 'CX', 'CY', 'CZ', 'DA', 'DB', 'DC', 'DD', 'DE', 'DF', 'DG', 'DH',  'DJ', 'DK', 'DL', 'DM', 'DN', 'DO', 'DQ', 'DR', 'DS', 'DT', 'DU', 'DV', 'DW', 'DX', 'DY', 'DZ', 'EA', 'EB', 'EC', 'ED', 'EE','EG', 'EH', 'EI', 'EJ', 'EK', 'EL']

    for column in columns_to_process:
        # 指定された列を選択
        col_range = sheet.Range(f"{column}1:{column}1048576")  
        col_range.TextToColumns(
            Destination=col_range.Cells(1, 1),  # データを上書きして貼り付け
            DataType=1,  # 1 は区切り文字でデータを区切る
            TextQualifier=1,  # ダブルクオートを文字列の区切りとする
            ConsecutiveDelimiter=True,  # 連続する区切り文字を1つの区切りとして扱う
            Tab=True,  # タブ区切りを有効にする（実際には自動的に検出される）
            Semicolon=False,
            Comma=True  # コンマ区切りを有効にする
        )

    # 保存して閉じる
    workbook.Save()
    workbook.Close()


####メイン処理####
# リストを作成
syori_taisyo = ["承認（不備ある）一覧", "差戻し一覧", "書面画像一覧","書面審査一覧"]
#syori_taisyo = ["書面審査一覧"]

# 引数の取得
mode = sys.argv[1]
date_mode=''
print(mode)

#書面ディレクトリ名作成
current_syomen_report_dir = syomen_report_dir + common_get_YYYYMMDD()[:8]
common_create_directory(current_syomen_report_dir)

# 引数 'w'：週次、 'm'：月次
if mode == "w":
    date_range = common_get_previous_week_date_range()
    date_mode='sql_w'
elif mode == "m":
    date_range = common_get_previous_month_date_range()
    date_mode='sql_m'


for i in range(len(syori_taisyo)):

    # SQLファイルパスを取得
    sql_file_path = common_get_file_path(current_dir,date_mode,syori_taisyo[i]+'.sql')

    # SQL実行結果を取得
    data = common_get_data("syomen",sql_file_path)

    # 既存のExcelファイル（ひな形）のパス
    excel_template_path = common_get_file_path(current_dir,'org','org.'+syori_taisyo[i]+'.xlsx') 

    # 出力先ファイルのパス
    output_excel_path   = common_get_file_path(current_dir,'',syori_taisyo[i] + date_range + '.xlsx') 

    # Excelテンプレート（ひな形）の読み込み
    workbook = load_workbook(excel_template_path)
    sheet = workbook.active

    # start=2で2行目から書き込み開始
    for j, row in enumerate(data, start=2):  
        # start=1で1列目から書き込み開始
        for k, value in enumerate(row, start=1):  
            # 「書面審査」のAV列以降の場合、標準で張り付ける
            if k>47 and syori_taisyo[i] =="書面審査一覧":  
    
                sheet.cell(row=j, column=k, value=value)
                continue

            # 先頭のゼロを保持したい場合、数値を文字列として書き込む
            if isinstance(value, int):  # 数値である場合
                # 文字列にしてゼロ埋め
                value = str(value).zfill(len(str(value)))  
            elif isinstance(value, str):  # 文字列である場合
                # 文字列としてそのまま
                pass

            sheet.cell(row=j, column=k, value=value)

    # 出力ファイルに保存
    workbook.save(output_excel_path)

    # 「書面審査」の場合、列の表示形式を変換する
    if syori_taisyo[i] =="書面審査一覧":  
        process_excel_columns(output_excel_path)
    
    # 作成したファイルを格納ディレクトリにコピー
    common_move_file(output_excel_path,current_syomen_report_dir)
